"""
Mind the Context — Streamlit front-end.
Run with:  streamlit run app.py
"""

from __future__ import annotations

import os
import tempfile

import pandas as pd
import streamlit as st
from dotenv import load_dotenv

import pipeline as P

load_dotenv()


# --- Small helpers -------------------------------------------------------
def _clean(v):
    """Return a stripped string or None for NaN/empty cells."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    return s or None


def _to_int(v):
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _save_tmp(uploaded) -> str:
    """Persist an uploaded file to a temp path so pipeline file-readers can open it."""
    suffix = os.path.splitext(uploaded.name)[1]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(uploaded.getbuffer())
    tmp.flush()
    tmp.close()
    return tmp.name


def _secret(name: str, default: str = "") -> str:
    """Read a value from Streamlit secrets if present (used on Streamlit Community Cloud),
    without crashing locally when no secrets.toml exists."""
    try:
        return str(st.secrets.get(name, default) or default)
    except Exception:
        return default


LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "logo.png")
_have_logo = os.path.exists(LOGO_PATH)

st.set_page_config(
    page_title="Context Minder",
    page_icon=LOGO_PATH if _have_logo else "🎯",
    layout="wide",
)

# --- Header --------------------------------------------------------------
hc1, hc2 = st.columns([1, 9], vertical_alignment="center")
with hc1:
    if _have_logo:
        st.image(LOGO_PATH, width=68)
with hc2:
    st.title("Context Minder")
st.markdown(
    "<p style='font-size:1.12rem; line-height:1.55; color:#3a3f45; margin-top:-0.4rem;'>"
    "Context-aware localization QA. It reads the <b>string ID, category, length, tags, "
    "speaker and comment</b> around each string — and the <b>style guide</b> and <b>glossary</b> "
    "— to flag translations that are wrong <i>for their context</i>, not just in isolation.</p>",
    unsafe_allow_html=True,
)
with st.expander("ℹ️ What gets checked — with vs without an LLM"):
    st.markdown(
        "**Always — no API key needed (deterministic rules):**\n"
        "- Max-length / UI overflow\n"
        "- Tag preservation — `[FR]`, `[u]`, `[one]`, `[ms]`/`[fs]`, etc.\n"
        "- Placeholder preservation — `{QUANTITY}`, `{gender}`, `{character.firstName}`…\n"
        "- Mob gender-variant pair — both `[ms]` and `[fs]` present\n"
        "- Empty / untranslated segments\n\n"
        "**With an LLM — needs an API key or a local model:**\n"
        "- Register & tone vs the speaker and category\n"
        "- Context fit vs the string ID + translator comment (e.g. question form vs button label)\n"
        "- Terminology / glossary adherence\n"
        "- Meaning fidelity & grammatical gender, using matched character / named-entity facts"
    )

# --- Sidebar: inputs -----------------------------------------------------
with st.sidebar:
    st.header("1 · Strings")
    xlsx_file = st.file_uploader("Strings file (.xlsx / .csv / .tsv)", type=["xlsx", "csv", "tsv"])

    st.header("2 · Context assets (optional)")
    glossary_file = st.file_uploader("Glossary (.csv)", type=["csv"])
    style_file = st.file_uploader("Style guide (.docx / .md)", type=["docx", "md", "markdown", "txt"])
    char_file = st.file_uploader(
        "Character descriptions (.md)",
        type=["md", "markdown", "txt"],
        help="Recommended for dialogue strings — personality/voice notes matched to each Speaker.",
    )
    st.caption("Character descriptions — recommended for dialogue strings.")
    with st.expander("…or paste character descriptions"):
        char_paste = st.text_area(
            "Paste character fragments",
            height=160,
            label_visibility="collapsed",
            placeholder=(
                "Paste from anywhere (Word, PDF, wiki). Any of these work:\n"
                "  ### Captain Vale\n  terse, commanding, never jokes\n\n"
                "  Mara: warm, witty, speaks formally to strangers\n"
                "  The Merchant — smooth, persuasive, slightly sly"
            ),
        )

    st.header("3 · Engine")
    provider = st.radio(
        "Provider",
        ["OpenAI (cloud)", "Local (Ollama / LM Studio)"],
        help="Local runs entirely on your machine — no data leaves it, no API cost. "
             "Expect weaker results than cloud models on the harder context/register checks.",
    )
    is_local = provider.startswith("Local")

    if is_local:
        base_url = st.text_input(
            "Local server URL",
            value=os.getenv("LOCAL_BASE_URL", "http://localhost:11434/v1"),
            help="Ollama: http://localhost:11434/v1  ·  LM Studio: http://localhost:1234/v1",
        )
        effective_key = ""  # not needed locally
        model = st.text_input("Model", value=os.getenv("LOCAL_MODEL", "llama3.1"))
        st.caption("Make sure the server is running and the model is pulled (e.g. `ollama run llama3.1`).")
        st.warning("Local mode only works when you run this app on your **own machine** — it connects to a "
                   "server on `localhost`. On the public hosted version, use OpenAI (cloud) or the deterministic checks.")
    else:
        base_url = ""
        api_key = os.getenv("OPENAI_API_KEY", "") or _secret("OPENAI_API_KEY")
        key_input = st.text_input(
            "OpenAI API key",
            value="",
            type="password",
            placeholder="sk-…",
            help="Paste your own OpenAI key to turn on the LLM checks. It is used only for this "
                 "session and is not stored or logged. If you run the app from your own copy of the "
                 "repo, you can instead keep the key in a local .env file and leave this blank. "
                 "With no key at all, the deterministic checks still run.",
        )
        effective_key = key_input.strip() or api_key
        if not effective_key:
            st.caption("No key yet — paste one above to enable LLM checks, or run deterministic-only.")
        model = st.text_input("Model", value=os.getenv("MIND_THE_CONTEXT_MODEL", "") or _secret("MIND_THE_CONTEXT_MODEL", "gpt-4o"))

    has_engine = is_local or bool(effective_key)
    use_llm = st.toggle("Use LLM checks", value=has_engine)
    if use_llm and not has_engine:
        st.warning("No engine configured — LLM checks will be skipped. Deterministic checks still run.")

    run = st.button("▶ Run QA", type="primary", use_container_width=True)


# --- File loading & column mapping --------------------------------------
# Recognised language codes / names -> a clean display label. Used both to
# auto-detect translation columns and to label them in the results.
LANG_NAMES = {
    "en": "EN", "eng": "EN", "english": "EN", "en-us": "EN", "en_us": "EN", "en-gb": "EN",
    "de": "DE", "ger": "DE", "deu": "DE", "german": "DE",
    "pl": "PL", "pol": "PL", "polish": "PL",
    "ru": "RU", "rus": "RU", "russian": "RU",
    "es": "ES", "spa": "ES", "spanish": "ES", "es-es": "ES", "es-419": "ES-419",
    "it": "IT", "ita": "IT", "italian": "IT",
    "fr": "FR", "fra": "FR", "fre": "FR", "french": "FR", "fr-fr": "FR", "fr-ca": "FR-CA",
    "zh": "ZH", "zh-cn": "ZH", "zh-hans": "ZH", "zh-tw": "ZH-TW", "chinese": "ZH",
    "ko": "KO", "kor": "KO", "korean": "KO",
    "ja": "JA", "jp": "JA", "jpn": "JA", "japanese": "JA",
    "pt": "PT", "pt-br": "PT-BR", "pt_br": "PT-BR", "ptbr": "PT-BR", "portuguese": "PT-BR", "pt-pt": "PT",
    "nl": "NL", "dutch": "NL", "tr": "TR", "turkish": "TR", "ar": "AR", "arabic": "AR",
    "cs": "CS", "czech": "CS", "sv": "SV", "swedish": "SV", "uk": "UK", "ukrainian": "UK",
}

ID_SYN = {"key", "id", "string id", "stringid", "string_id", "string-id", "identifier",
          "segment id", "segmentid", "resname", "name", "string key", "msgid"}
SRC_SYN = {"source", "source (en)", "source_en", "source-en", "src", "source text",
           "sourcetext", "en source", "original"}
SRC_LANG = {"en", "eng", "english", "en-us", "en_us", "en-gb"}
SPEAKER_SYN = {"speaker", "character", "char", "who", "actor", "voice", "persona"}
CATEGORY_SYN = {"category", "type", "content type", "contenttype", "context", "kind", "scene"}
COMMENT_SYN = {"comment", "comments", "note", "notes", "developer notes", "dev notes", "devnote",
               "description", "instructions", "instruction", "remark", "context note", "devcomment"}
MAXLEN_SYN = {"max length", "maxlength", "max_len", "maxlen", "max length (chars)", "limit",
              "char limit", "charlimit", "length limit", "maxchars", "max chars"}
TAGS_SYN = {"tags", "tag", "flags", "flag", "markers"}


def load_strings(file) -> pd.DataFrame:
    """Read the strings table from xlsx/csv/tsv. For xlsx, prefer a 'Strings' sheet,
    otherwise use the first sheet — never assume a fixed sheet name."""
    name = (getattr(file, "name", "") or "").lower()
    if name.endswith(".csv"):
        df = pd.read_csv(file)
    elif name.endswith(".tsv") or name.endswith(".tab"):
        df = pd.read_csv(file, sep="\t")
    else:
        xls = pd.ExcelFile(file)
        sheet = "Strings" if "Strings" in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def lang_label(col: str) -> str:
    """Clean display label for a translation column ('ger' -> 'DE', 'Target (PL)' -> 'PL')."""
    s = str(col).strip()
    low = s.lower()
    if low in LANG_NAMES:
        return LANG_NAMES[low]
    if "(" in s and ")" in s:                       # 'Target (PL)'
        inside = s[s.find("(") + 1:s.find(")")].strip()
        return inside.upper() or s.upper()
    return s.upper()


def _find(cols, low, cands):
    for c in cols:
        if low[c] in cands:
            return c
    return None


def auto_map(cols) -> dict:
    """Best-guess mapping from arbitrary column names to our roles."""
    cols = list(cols)
    low = {c: str(c).strip().lower() for c in cols}
    id_col = _find(cols, low, ID_SYN)
    source_col = _find(cols, low, SRC_SYN) or _find(cols, low, SRC_LANG)
    speaker = _find(cols, low, SPEAKER_SYN)
    category = _find(cols, low, CATEGORY_SYN)
    comment = _find(cols, low, COMMENT_SYN)
    maxlen = _find(cols, low, MAXLEN_SYN)
    tags = _find(cols, low, TAGS_SYN)
    meta = {id_col, source_col, speaker, category, comment, maxlen, tags}
    targets = [c for c in cols
               if c not in meta and (low[c] in LANG_NAMES or low[c].startswith("target"))]
    return {"id": id_col, "source": source_col, "targets": targets, "speaker": speaker,
            "category": category, "comment": comment, "maxlen": maxlen, "tags": tags}


def _safe_idx(options, value, default=0):
    return options.index(value) if value in options else default


def column_mapping_ui(df: pd.DataFrame) -> dict:
    """Render the mapping panel (pre-filled with auto-detection) and return the mapping."""
    cols = list(df.columns)
    guess = auto_map(cols)
    detected_ok = bool(guess["id"] and guess["source"] and guess["targets"])
    NONE = "— none —"

    label = "🔧 Column mapping" + ("" if detected_ok else "  ⚠ please review")
    with st.expander(label, expanded=not detected_ok):
        if detected_ok:
            st.caption(f"Auto-detected from your headers ({len(cols)} columns). Adjust anything that looks wrong.")
        else:
            st.caption("Couldn't fully auto-detect the columns — please set them below.")

        c1, c2 = st.columns(2)
        id_col = c1.selectbox("String ID column", cols, index=_safe_idx(cols, guess["id"]), key="map_id")
        source_col = c2.selectbox("Source (original) column", cols, index=_safe_idx(cols, guess["source"]), key="map_src")
        targets = st.multiselect(
            "Translation columns to check",
            [c for c in cols if c != source_col],
            default=[c for c in guess["targets"] if c != source_col],
            key="map_targets",
            help="Each selected column is checked as a separate target language.",
        )
        st.markdown("**Context columns (optional)** — improve the context-aware checks.")
        d1, d2, d3, d4, d5 = st.columns(5)
        opt = [NONE] + cols
        speaker = d1.selectbox("Speaker", opt, index=_safe_idx(opt, guess["speaker"]), key="map_speaker")
        category = d2.selectbox("Category", opt, index=_safe_idx(opt, guess["category"]), key="map_category")
        comment = d3.selectbox("Comment", opt, index=_safe_idx(opt, guess["comment"]), key="map_comment")
        maxlen = d4.selectbox("Max length", opt, index=_safe_idx(opt, guess["maxlen"]), key="map_maxlen")
        tags = d5.selectbox("Tags", opt, index=_safe_idx(opt, guess["tags"]), key="map_tags")

    def nn(v):
        return None if v == NONE else v

    return {"id": id_col, "source": source_col, "targets": targets, "speaker": nn(speaker),
            "category": nn(category), "comment": nn(comment), "maxlen": nn(maxlen), "tags": nn(tags)}


def build_xlsx_report(results) -> bytes:
    """Build an XLSX QA report whose columns map 1:1 onto the memoQ feedback pattern:
    Status = segment status (Rejected / Review / OK), Comment = note for the memoQ comment field."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    status_map = {"error": "Rejected", "warning": "Rejected", "suggestion": "Review", "ok": "OK"}
    red = PatternFill("solid", fgColor="F4CCCC")
    amber = PatternFill("solid", fgColor="FCE5CD")

    wb = Workbook()
    ws = wb.active
    ws.title = "QA report"
    headers = ["String ID", "Source (EN)", "Target lang", "Translation",
               "Status", "Severity", "Issue type", "Comment", "Suggestion"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for r in results:
        issues = r["issues"]
        status = status_map.get(r["status"], r["status"])
        types = ", ".join(sorted({i["type"] for i in issues}))
        comment = " | ".join(i["message"] for i in issues)
        suggestion = " | ".join(i["suggestion"] for i in issues if i["suggestion"])
        ws.append([r["string_id"], r["source"], r["language"], r["target"],
                   status, r["status"], types, comment, suggestion])
        fill = red if status == "Rejected" else (amber if status == "Review" else None)
        if fill:
            for c in ws[ws.max_row]:
                c.fill = fill

    widths = [22, 42, 11, 42, 10, 12, 18, 50, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for col in ("B", "D", "H", "I"):
        for cell in ws[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def make_client(key: str, base_url: str = ""):
    """Build an OpenAI-compatible client. A base_url points it at a local server
    (Ollama / LM Studio), which speak the same API, so no extra dependency is needed."""
    try:
        import openai
        if base_url:
            # local servers don't need a real key, but the SDK requires a non-empty string
            return openai.OpenAI(api_key=key or "local", base_url=base_url)
        if not key:
            return None
        return openai.OpenAI(api_key=key)
    except Exception as e:
        st.error(f"Could not init client: {e}")
        return None


STATUS_STYLE = {
    "error": ("🔴", "error"),
    "warning": ("🟠", "warning"),
    "suggestion": ("🟡", "suggestion"),
    "ok": ("🟢", "ok"),
}


def language_selector(df: pd.DataFrame, target_cols: list, use_llm: bool) -> list:
    """Prominent 'Languages to analyze' picker — lets the user restrict the run to a
    subset of detected languages to save time and tokens. Returns the chosen columns."""
    label_to_col, labels = {}, []
    for c in target_cols:
        lab = lang_label(c)
        if lab in label_to_col:          # disambiguate duplicate labels
            lab = f"{lab} ({c})"
        label_to_col[lab] = c
        labels.append(lab)

    st.subheader("Languages to analyze")
    chosen = st.multiselect(
        "Pick the languages to run — fewer languages = faster and cheaper.",
        labels, default=labels, key="analyze_langs",
    )
    cols = [label_to_col[l] for l in chosen]
    n_rows, n_lang = len(df), len(cols)
    segments = n_rows * n_lang
    if not cols:
        st.warning("No language selected — pick at least one to run.")
    elif use_llm:
        st.caption(f"Workload: **{n_rows} rows × {n_lang} languages = {segments} segments**, "
                   f"up to **{segments} LLM calls**. Token usage is reported after the run.")
    else:
        st.caption(f"Workload: **{segments} segments** (deterministic checks only — no API tokens).")
    return cols


# --- Load file + show mapping (whenever a file is present) ----------------
df = None
mapping = None
analyze_cols = []
if xlsx_file is not None:
    try:
        df = load_strings(xlsx_file)
    except Exception as e:
        st.error(f"Could not read the file: {e}")
        df = None
    if df is not None:
        if df.empty:
            st.warning("The file has no data rows.")
        mapping = column_mapping_ui(df)
        if mapping and mapping["targets"]:
            analyze_cols = language_selector(df, mapping["targets"], use_llm)


# --- Run -----------------------------------------------------------------
if run:
    if df is None:
        st.error("Upload a strings file first.")
        st.stop()
    if not mapping or not mapping["targets"]:
        st.error("Pick at least one translation column in **Column mapping** above.")
        st.stop()
    tgt_cols = analyze_cols or mapping["targets"]
    if not tgt_cols:
        st.error("Select at least one language in **Languages to analyze** above.")
        st.stop()

    glossary = P.load_glossary(_save_tmp(glossary_file)) if glossary_file else []
    style_sections = P.load_style_guide(_save_tmp(style_file)) if style_file else {}
    characters = P.load_character_bible(_save_tmp(char_file)) if char_file else []
    if char_paste and char_paste.strip():
        characters = characters + P.parse_character_text(char_paste)

    m = mapping
    client = make_client(effective_key, base_url) if use_llm else None

    results = []
    total = len(df) * len(tgt_cols)
    bar = st.progress(0.0, text="Analyzing…")
    done = 0
    for _, r in df.iterrows():
        for col in tgt_cols:
            row = P.StringRow(
                string_id=str(r.get(m["id"], "") or "") if m["id"] else "",
                source=str(r.get(m["source"], "") or "") if m["source"] else "",
                target=str(r.get(col, "") or ""),
                language=lang_label(col),
                speaker=_clean(r.get(m["speaker"])) if m["speaker"] else None,
                category=_clean(r.get(m["category"])) if m["category"] else None,
                comment=_clean(r.get(m["comment"])) if m["comment"] else None,
                max_length=_to_int(r.get(m["maxlen"])) if m["maxlen"] else None,
                tags=_clean(r.get(m["tags"])) if m["tags"] else None,
            )
            results.append(P.analyze_row(row, glossary, style_sections, client, model, use_llm, characters))
            done += 1
            bar.progress(done / total, text=f"Analyzing… {done}/{total}")
    bar.empty()
    st.session_state["results"] = results
    st.session_state["glossary_n"] = len(glossary)
    st.session_state["style_n"] = len(style_sections)
    st.session_state["char_n"] = len(characters)
    st.session_state["tokens"] = {
        "prompt": sum(r["tokens"]["prompt"] for r in results),
        "completion": sum(r["tokens"]["completion"] for r in results),
        "calls": sum(r["tokens"]["calls"] for r in results),
    }


# --- Results -------------------------------------------------------------
results = st.session_state.get("results")
if results:
    n_err = sum(1 for r in results if r["status"] == "error")
    n_warn = sum(1 for r in results if r["status"] == "warning")
    n_sug = sum(1 for r in results if r["status"] == "suggestion")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Strings checked", len(results))
    c2.metric("🔴 Errors", n_err)
    c3.metric("🟠 Warnings", n_warn)
    c4.metric("🟡 Suggestions", n_sug)
    c5.metric("🟢 Clean", sum(1 for r in results if r["status"] == "ok"))

    st.caption(
        f"Context loaded: {st.session_state.get('glossary_n', 0)} glossary terms · "
        f"{st.session_state.get('style_n', 0)} style-guide sections · "
        f"{st.session_state.get('char_n', 0)} character profiles."
    )

    tk = st.session_state.get("tokens")
    if tk and tk.get("calls"):
        t1, t2, t3 = st.columns(3)
        t1.metric("⬆ Tokens sent", f"{tk['prompt']:,}")
        t2.metric("⬇ Tokens received", f"{tk['completion']:,}")
        t3.metric("LLM calls", f"{tk['calls']:,}")
        st.caption("Tokens drive API cost. Trim **Languages to analyze** to spend less. "
                   "(Local models report 0 if the server doesn't return usage.)")

    f1, f2 = st.columns([1, 1])
    show = f1.multiselect(
        "Show severities",
        ["error", "warning", "suggestion", "ok"],
        default=["error", "warning", "suggestion"],
    )
    langs = sorted({r["language"] for r in results})
    lang_filter = f2.multiselect("Languages", langs, default=langs)

    view = [r for r in results if r["status"] in show and r["language"] in lang_filter]

    # Summary table
    table = pd.DataFrame(
        [
            {
                "": STATUS_STYLE.get(r["status"], ("", ""))[0],
                "String ID": r["string_id"],
                "Lang": r["language"],
                "Issues": r["n_issues"],
                "Source": (r["source"][:60] + "…") if len(r["source"]) > 60 else r["source"],
                "Translation": (r["target"][:60] + "…") if len(r["target"]) > 60 else r["target"],
            }
            for r in view
        ]
    )
    st.dataframe(table, use_container_width=True, hide_index=True)

    # Detail cards
    st.subheader("Details")
    for r in view:
        if r["status"] == "ok":
            continue
        icon = STATUS_STYLE.get(r["status"], ("", ""))[0]
        with st.expander(f"{icon} {r['string_id']} · {r['language']} · {r['n_issues']} issue(s)"):
            st.markdown(f"**Source:** {r['source']}")
            st.markdown(f"**Translation:** {r['target'] or '_(empty)_'}")
            for iss in r["issues"]:
                tag = {"rule": "⚙️ rule", "llm": "🧠 llm"}.get(iss["source"], "")
                st.markdown(
                    f"- **[{iss['severity']}] {iss['type']}** {tag} — {iss['message']}"
                    + (f"  \n  ↳ _{iss['suggestion']}_" if iss["suggestion"] else "")
                )
            ctx = r["context"]
            if ctx["glossary"] or ctx["style_sections"] or ctx["inferred"] or ctx.get("character") or ctx.get("entities"):
                with st.popover("Context used"):
                    if ctx["inferred"]:
                        st.markdown("**Inferred from ID/metadata**")
                        for n in ctx["inferred"]:
                            st.markdown(f"- {n}")
                    if ctx["glossary"]:
                        st.markdown("**Glossary matches**")
                        for g in ctx["glossary"]:
                            st.markdown(f"- {g['english']} → {g['term']}")
                    if ctx["style_sections"]:
                        st.markdown("**Style-guide sections pulled**")
                        for s in ctx["style_sections"]:
                            st.code(s[:600])
                    if ctx.get("character"):
                        st.markdown(f"**Speaker profile — {ctx['character']['name']}**")
                        st.caption(ctx["character"]["voice"][:400])
                    if ctx.get("entities"):
                        st.markdown("**Named entities in source**")
                        for e in ctx["entities"]:
                            st.caption(f"{e['name']}: {e['voice'][:200]}")

    # Export — XLSX mapped onto the memoQ feedback pattern (segment status + comment)
    st.download_button(
        "⬇ Download QA report (XLSX)",
        build_xlsx_report(results),
        file_name="qa_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption("Columns map 1:1 onto memoQ: **Status** = segment status (Rejected / OK), "
               "**Comment** = the note for the memoQ comment field.")
else:
    st.info("Upload a strings file in the sidebar, check the **Column mapping**, then click **Run QA**. "
            "Add the glossary, style guide and character descriptions for the full context-aware checks.")
